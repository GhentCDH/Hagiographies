# cli.py
# ---------------------------------------------------------------------------
# Import the June 2026 corpus workbook into PostgreSQL.
#
# Source sheets (resolved by stripped name — ' MANUSCRIPTS' has a leading
# space in the workbook):
#   'TEXTS'        → Text (+ Author, Place, ChurchEntity, Typology)
#   'MANUSCRIPTS'  → Codex + Manuscript copies (+ Image, relations)
#   'EDITIONS'     → Edition (+ EditionVolume, EditionManuscript, scans)
#
# Layout notes that drive the parsing:
#   - All three sheets have a real header row (row 1).  Every expected header
#     is verified up-front; a missing header aborts the import loudly.
#   - GPS is encoded differently per sheet: the TEXTS sheet stores unscaled
#     integers (49119308 == 49.119308) so they are divided by 1e6; the
#     MANUSCRIPTS sheet already stores decimal degrees, used verbatim.
#     The 'GPS Longitude' columns hold latitude (~49 for Metz) and vice
#     versa, so lon/lat are swapped on read.
#   - One Codex row per 'Codex unique identifier'; one Manuscript copy row
#     per sheet row, keyed on 'Manuscript copy unique identifier per text'
#     (e.g. "29-1") — the value editions and exemplar relations reference.
#   - Unresolvable cross-sheet references are recorded in the import report
#     under 'link_unresolved' instead of failing the import.
# ---------------------------------------------------------------------------

import logging
import re
import unicodedata
from itertools import count, islice
from typing import Any, Dict, Generator, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet
from rich.logging import RichHandler
from sqlmodel import SQLModel, Session, select

from utilities.config import EXCEL, DATA_ROOT
from utilities.db import engine
from utilities.model import (
    Author,
    Certainty,
    ChurchEntity,
    Codex,
    DatingCentury,
    Edition,
    EditionConsultedVolume,
    EditionExternalResource,
    EditionManuscript,
    EditionVolume,
    ExternalResource,
    ExternalResourceType,
    Image,
    ImageType,
    Institution,
    Manuscript,
    ManuscriptRelation,
    ManuscriptType,
    Place,
    RelationType,
    Text,
    Typology,
    VernacularRegion,
)

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------

handler = RichHandler(
    rich_tracebacks=True,
    tracebacks_show_locals=True,
    markup=True,
    show_time=True,
    show_path=True,
)
logging.basicConfig(level=logging.INFO, format="%(message)s", handlers=[handler])
logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

EMPTY_ROW_LIMIT = 10

SHEET_TEXTS = "TEXTS"
SHEET_MANUSCRIPTS = "MANUSCRIPTS"
SHEET_EDITIONS = "EDITIONS"

# Author name variants treated as anonymous — each occurrence becomes a
# distinct Author record (Anon.-1, Anon.-2, …) so per-text metadata survives.
_ANON_NAMES = {"anon", "anon.", "anonymous"}

# Valid range for 'Manuscript dating by (earliest) century'.  Date-typed cells
# would otherwise leak a year (e.g. 2026) through parse_int.
_CENTURY_MIN, _CENTURY_MAX = 1, 21

# Values (case-insensitive) treated as "no value" for any cell.
_NULL_TOKENS = {"", "n/a", "na", "none", "unknown", "nan", "-"}

# Image-type markers, longest first so "IIIF Microfilm" wins over "IIIF".
_IMAGE_TYPE_MAP = [
    ("IIIF MICROFILM", "iiif_mf"),
    ("IIIF MF", "iiif_mf"),
    ("IIIF", "iiif"),
    ("SCAN", "scan"),
    ("IPHONE", "iphone_photo"),
    ("PHOTO", "iphone_photo"),
]

URL_RE = re.compile(r"^https?://[^\s/$.?#].[^\s]*$", re.IGNORECASE)


# ---------------------------------------------------------------------------
# Pure value helpers
# ---------------------------------------------------------------------------

def clean_value(val: Any) -> Optional[str]:
    """Return a stripped string, or None for empty / placeholder values."""
    if val is None:
        return None
    s = str(val).strip()
    if s.lower() in _NULL_TOKENS:
        return None
    return s or None


def parse_int(val: Any) -> Optional[int]:
    """Parse an integer (tolerating floats and '1047/48'-style ranges)."""
    s = clean_value(val)
    if s is None:
        return None
    # Take the leading integer token ("1047/48" → 1047, "c. 1100" → 1100).
    m = re.search(r"-?\d+", s.replace(",", ""))
    if not m:
        return None
    try:
        return int(m.group(0))
    except ValueError:
        return None


def parse_float(val: Any) -> Optional[float]:
    """Parse a float, returning None on failure or placeholder."""
    s = clean_value(val)
    if s is None:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def parse_gps_int(val: Any) -> Optional[float]:
    """Parse an unscaled-integer GPS value (49119308 → 49.119308).

    The TEXTS sheet stores coordinates as integers with six implied decimals.
    Already-decimal values (rare) are detected and returned as-is.
    """
    f = parse_float(val)
    if f is None:
        return None
    if abs(f) > 180:  # integer-encoded
        return f / 1_000_000
    return f


def parse_yesno(val: Any) -> Optional[bool]:
    """Convert a yes/no Excel value to a bool; None for blank/unrecognised."""
    s = clean_value(val)
    if s is None:
        return None
    u = s.upper()
    if u in ("Y", "YES", "1", "TRUE", "OUI", "JA"):
        return True
    if u in ("N", "NO", "0", "FALSE", "NON", "NEE"):
        return False
    return None


def _normalize_name(value: Optional[str]) -> Optional[str]:
    """Normalise a place/institution/author name: lowercase, collapse ws, fold accents."""
    s = clean_value(value)
    if s is None:
        return None
    s = " ".join(s.split()).lower()
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
    return s or None


def _confidence_value(
    val: Any, report: "ImportReport", excel_row: int, col: str
) -> Optional[str]:
    """Return a confidence-rating cell value, rejecting bare numbers.

    A numeric value in a confidence column indicates a shifted source row
    (a GPS coordinate leaking in) — store NULL and report it instead.
    """
    s = clean_value(val)
    if s is None:
        return None
    if _FLOAT_RE.match(s):
        report.add("invalid_format", {
            "Row": excel_row,
            "Column": col,
            "Value": s,
            "Reason": "Numeric value in a confidence column — "
                      "shifted cells in the source row?",
        })
        return None
    return s


def _clean_uncertain(value: Optional[str]) -> Tuple[Optional[str], bool]:
    """Strip '?' markers from a name, returning (cleaned name, was_uncertain).

    'Metz (?)' → ('Metz', True); the uncertainty moves into a confidence
    rating column instead of polluting the name (researcher feedback).
    """
    s = clean_value(value)
    if s is None:
        return None, False
    uncertain = "?" in s
    if uncertain:
        s = s.replace("?", "")
        s = re.sub(r"\(\s*\)", "", s)  # drop now-empty parens
        s = " ".join(s.split()).strip(" ,;")
    return (s or None), uncertain


def _normalize_vol(value: Any) -> Optional[str]:
    """Normalise a volume identifier for matching.

    Unlike _normalize_ref, parentheticals are significant here
    ('AASS Jul. 4 (1st.)' vs '(3rd.)'), so only whitespace and case are
    folded: 'AASS OSB 2 (1 st.)' == 'AASS OSB 2 (1st.)'.
    """
    s = clean_value(value)
    if s is None:
        return None
    return re.sub(r"\s+", "", s).lower()


def _normalize_ref(value: Any) -> Optional[str]:
    """Normalise a cross-sheet reference key (e.g. '29-1') for matching.

    Trailing parenthetical qualifiers ('3621-4 (?)') are stripped so the bare
    identifier still resolves; genuinely compound refs ('8709-3 or 8709-5')
    are left intact and will be reported as unresolved.
    """
    s = clean_value(value)
    if s is None:
        return None
    s = re.sub(r"\([^)]*\)", "", s)  # drop parenthetical qualifiers
    return re.sub(r"\s+", "", s).lower()


def _is_empty_row(cells: List[Cell]) -> bool:
    return all(c.value is None or str(c.value).strip() == "" for c in cells)


def _iter_data_rows(
    rows_iter, sheet_title: str = "", empty_limit: int = EMPTY_ROW_LIMIT
) -> Generator[Tuple[int, List[Cell]], None, None]:
    """Yield (1-based row number, cell list), stopping after empty_limit blanks."""
    consecutive_empty = 0
    for row_num, row_cells in enumerate(rows_iter, start=2):
        cells = list(row_cells)
        if _is_empty_row(cells):
            consecutive_empty += 1
            if consecutive_empty >= empty_limit:
                logger.info(
                    f"[{sheet_title}] {empty_limit} consecutive empty rows "
                    f"at row {row_num} — stopping early."
                )
                return
            continue
        consecutive_empty = 0
        yield row_num, cells


def _extract_hyperlink_url(cell: Optional[Cell]) -> Optional[str]:
    """Extract a URL from a cell's hyperlink target (or value), else None.

    Cells whose display text is the literal "Link" carry the real URL in the
    hyperlink target; cells with no usable target yield None.
    """
    if cell is None:
        return None
    if cell.hyperlink:
        target = (
            getattr(cell.hyperlink, "target", None)
            or getattr(cell.hyperlink, "location", None)
        )
        if target and URL_RE.match(str(target).strip()):
            return str(target).strip()
    if isinstance(cell.value, str) and URL_RE.match(cell.value.strip()):
        return cell.value.strip()
    return None


def _infer_image_type(image_availability: Optional[str]) -> str:
    aa = (image_availability or "").upper()
    for marker, itype in _IMAGE_TYPE_MAP:
        if marker in aa:
            return itype
    return "scan"


def _chunked(iterable, n: int):
    it = iter(iterable)
    while True:
        chunk = list(islice(it, n))
        if not chunk:
            return
        yield chunk


# ---------------------------------------------------------------------------
# Header-indexed row access
# ---------------------------------------------------------------------------

def _normalize_col(col: str) -> str:
    """Collapse whitespace in a header lookup key."""
    return re.sub(r"\s+", " ", col.strip())


def _get_sheet(wb, name: str) -> Worksheet:
    """Resolve a worksheet by stripped title (' MANUSCRIPTS' has a leading space)."""
    for title in wb.sheetnames:
        if title.strip() == name:
            return wb[title]
    raise KeyError(f"Sheet {name!r} not found; workbook has {wb.sheetnames}")


def _require_headers(
    headers: Dict[str, int], expected: List[str], sheet: str
) -> None:
    """Abort loudly if any expected header is missing from the sheet."""
    missing = [c for c in expected if _normalize_col(c) not in headers]
    if missing:
        raise ValueError(
            f"[{sheet}] missing expected column headers: {missing}\n"
            f"Headers found: {sorted(headers)}"
        )


def _read_headers(ws: Worksheet) -> Tuple[Dict[str, int], Any]:
    """Read row 1 as headers, returning {normalised header → column index}.

    Duplicate headers get suffixes (_2, _3, …) matching the order seen, so the
    repeated 'Links to repertories' / 'Manuscript used N' style columns remain
    individually addressable.
    """
    rows_iter = ws.rows
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return {}, iter([])

    index: Dict[str, int] = {}
    seen: Dict[str, int] = {}
    for i, c in enumerate(header_row):
        if c.value is None:
            continue
        name = _normalize_col(str(c.value))
        if not name:
            continue
        if name in seen:
            seen[name] += 1
            name = f"{name}_{seen[name]}"
        else:
            seen[name] = 1
        index[name] = i
    logger.info(f"[Headers] {list(index.keys())}")
    return index, rows_iter


class Row:
    """Header-keyed accessor over one worksheet data row."""

    def __init__(self, cells: List[Cell], header_index: Dict[str, int]) -> None:
        self._cells = cells
        self._index = header_index

    def cell(self, col: str) -> Optional[Cell]:
        idx = self._index.get(_normalize_col(col))
        if idx is None or idx >= len(self._cells):
            return None
        return self._cells[idx]

    def raw(self, col: str) -> Any:
        c = self.cell(col)
        return c.value if c else None

    def s(self, col: str) -> Optional[str]:
        return clean_value(self.raw(col))

    def i(self, col: str) -> Optional[int]:
        return parse_int(self.raw(col))

    def f(self, col: str) -> Optional[float]:
        return parse_float(self.raw(col))

    def yesno(self, col: str) -> Optional[bool]:
        return parse_yesno(self.raw(col))


# ---------------------------------------------------------------------------
# Cached get_or_create helpers
# ---------------------------------------------------------------------------

def _get_or_create_place(
    session: Session,
    name: Optional[str],
    cache: Dict[str, Place],
    lat: Optional[float] = None,
    lon: Optional[float] = None,
) -> Optional[int]:
    raw, uncertain = _clean_uncertain(name)
    if raw is None:
        return None
    if lat is None or lon is None:  # a point needs both coordinates
        lat = lon = None
    norm = _normalize_name(raw)

    def _enrich(p: Place) -> None:
        if lat is not None and p.lat is None:
            p.lat = lat
        if lon is not None and p.lon is None:
            p.lon = lon
        if uncertain and p.confidence_rating is None:
            p.confidence_rating = "uncertain"

    if norm in cache:
        _enrich(cache[norm])
        return cache[norm].id
    matched = next(
        (p for p in session.exec(select(Place)).all()
         if _normalize_name(p.name) == norm),
        None,
    )
    if matched:
        _enrich(matched)
        cache[norm] = matched
        return matched.id
    place = Place(
        name=raw, lat=lat, lon=lon,
        confidence_rating="uncertain" if uncertain else None,
    )
    session.add(place)
    session.flush()
    cache[norm] = place
    return place.id


def _get_or_create_institution(
    session: Session,
    name: Optional[str],
    place_id: Optional[int],
    cache: Dict[str, Institution],
    lat: Optional[float] = None,
    lon: Optional[float] = None,
) -> Optional[int]:
    raw, uncertain = _clean_uncertain(name)
    if raw is None:
        return None
    if lat is None or lon is None:  # a point needs both coordinates
        lat = lon = None
    norm = _normalize_name(raw)

    def _enrich(i: Institution) -> None:
        if place_id and not i.place_id:
            i.place_id = place_id
        if lat is not None and i.lat is None:
            i.lat = lat
        if lon is not None and i.lon is None:
            i.lon = lon
        if uncertain and i.confidence_rating is None:
            i.confidence_rating = "uncertain"

    if norm in cache:
        _enrich(cache[norm])
        return cache[norm].id
    matched = next(
        (i for i in session.exec(select(Institution)).all()
         if _normalize_name(i.name) == norm),
        None,
    )
    if matched:
        _enrich(matched)
        cache[norm] = matched
        return matched.id
    inst = Institution(
        name=raw, place_id=place_id, lat=lat, lon=lon,
        confidence_rating="uncertain" if uncertain else None,
    )
    session.add(inst)
    session.flush()
    cache[norm] = inst
    return inst.id


def _get_or_create_author(
    session: Session,
    name: Optional[str],
    cache: Dict[str, Author],
    anon_seq,
    institutional_training_ground: Optional[str] = None,
    regional_antecedents: Optional[str] = None,
    milieu: Optional[str] = None,
) -> Optional[int]:
    raw = clean_value(name)
    if raw is None:
        return None
    norm = _normalize_name(raw)
    if norm is None:
        return None

    # Anonymous authors stay distinct: one Author per text (Anon.-1, Anon.-2, …)
    # so their per-text training ground / antecedents / milieu are not merged.
    if norm in _ANON_NAMES:
        auth = Author(
            name=f"Anon.-{next(anon_seq)}",
            institutional_training_ground=institutional_training_ground,
            regional_antecedents=regional_antecedents,
            milieu=milieu,
        )
        session.add(auth)
        session.flush()
        return auth.id

    def _backfill(a: Author) -> None:
        if institutional_training_ground and not a.institutional_training_ground:
            a.institutional_training_ground = institutional_training_ground
        if regional_antecedents and not a.regional_antecedents:
            a.regional_antecedents = regional_antecedents
        if milieu and not a.milieu:
            a.milieu = milieu

    if norm in cache:
        _backfill(cache[norm])
        return cache[norm].id
    matched = next(
        (a for a in session.exec(select(Author)).all()
         if _normalize_name(a.name) == norm),
        None,
    )
    if matched:
        _backfill(matched)
        cache[norm] = matched
        return matched.id
    auth = Author(
        name=raw,
        institutional_training_ground=institutional_training_ground,
        regional_antecedents=regional_antecedents,
        milieu=milieu,
    )
    session.add(auth)
    session.flush()
    cache[norm] = auth
    return auth.id


def _get_or_create_typology(
    session: Session,
    name: Optional[str],
    parent_id: Optional[int],
    cache: Dict[str, Typology],
) -> Optional[int]:
    raw = clean_value(name)
    if raw is None:
        return None
    if raw in cache:
        return cache[raw].id
    existing = session.exec(select(Typology).where(Typology.name == raw)).first()
    if existing:
        if parent_id and not existing.parent_id:
            existing.parent_id = parent_id
        cache[raw] = existing
        return existing.id
    typo = Typology(name=raw, parent_id=parent_id)
    session.add(typo)
    session.flush()
    cache[raw] = typo
    return typo.id


def _get_or_create_manuscript_type(
    session: Session, name: Optional[str], cache: Dict[str, ManuscriptType]
) -> Optional[int]:
    raw = clean_value(name)
    if raw is None:
        return None
    if raw in cache:
        return cache[raw].id
    existing = session.exec(
        select(ManuscriptType).where(ManuscriptType.name == raw)
    ).first()
    if existing:
        cache[raw] = existing
        return existing.id
    mt = ManuscriptType(name=raw)
    session.add(mt)
    session.flush()
    cache[raw] = mt
    return mt.id


def _get_or_create_image_type(
    session: Session, name: Optional[str], cache: Dict[str, ImageType]
) -> Optional[int]:
    raw = clean_value(name)
    if raw is None:
        return None
    if raw in cache:
        return cache[raw].id
    existing = session.exec(select(ImageType).where(ImageType.name == raw)).first()
    if existing:
        cache[raw] = existing
        return existing.id
    obj = ImageType(name=raw)
    session.add(obj)
    session.flush()
    cache[raw] = obj
    return obj.id


def _get_or_create_church_entity(
    session: Session,
    name: Optional[str],
    entity_type: str,
    cache: Dict[str, ChurchEntity],
    lat: Optional[float] = None,
    lon: Optional[float] = None,
) -> Optional[int]:
    """Get/create a ChurchEntity.

    lat/lon are the fallback GPS: when a row has no origin institution, the
    sheet's coordinates belong to the diocese and are stored here (backfilled
    only while the entity's coordinates are still NULL).
    """
    raw, uncertain = _clean_uncertain(name)
    if raw is None:
        return None
    if lat is None or lon is None:  # a point needs both coordinates
        lat = lon = None
    key = f"{raw}|{entity_type}"

    def _enrich(ce: ChurchEntity) -> None:
        if lat is not None and ce.lat is None:
            ce.lat = lat
        if lon is not None and ce.lon is None:
            ce.lon = lon
        if uncertain and ce.confidence_rating is None:
            ce.confidence_rating = "uncertain"

    if key in cache:
        _enrich(cache[key])
        return cache[key].id
    existing = session.exec(
        select(ChurchEntity).where(
            ChurchEntity.name == raw,
            ChurchEntity.entity_type == entity_type,
        )
    ).first()
    if existing:
        _enrich(existing)
        cache[key] = existing
        return existing.id
    ce = ChurchEntity(
        name=raw, entity_type=entity_type, lat=lat, lon=lon,
        confidence_rating="uncertain" if uncertain else None,
    )
    session.add(ce)
    session.flush()
    cache[key] = ce
    return ce.id


def _get_or_create_edition_volume(
    session: Session, identifier: Optional[str], cache: Dict[str, EditionVolume]
) -> Optional[EditionVolume]:
    raw = clean_value(identifier)
    norm = _normalize_vol(raw)
    if norm is None:
        return None
    if norm in cache:
        return cache[norm]
    existing = next(
        (v for v in session.exec(select(EditionVolume)).all()
         if _normalize_vol(v.identifier) == norm),
        None,
    )
    if existing:
        cache[norm] = existing
        return existing
    vol = EditionVolume(identifier=raw)
    session.add(vol)
    session.flush()
    cache[norm] = vol
    return vol


def _get_or_create_dating_century(
    session: Session,
    century_val: Any,
    cache: Dict[int, DatingCentury],
    report: Optional["ImportReport"] = None,
    excel_row: Optional[int] = None,
) -> Optional[int]:
    century = parse_int(century_val)
    if century is None:
        return None
    if not _CENTURY_MIN <= century <= _CENTURY_MAX:
        if report is not None:
            report.add("invalid_format", {
                "Row": excel_row or 0,
                "Column": "Manuscript dating by (earliest) century",
                "Value": str(century_val),
                "Reason": f"Century {century} outside {_CENTURY_MIN}–{_CENTURY_MAX} "
                          "(date-typed cell?) — stored as NULL",
            })
        return None
    if century in cache:
        return cache[century].id
    existing = session.exec(
        select(DatingCentury).where(DatingCentury.century == century)
    ).first()
    if existing:
        cache[century] = existing
        return existing.id
    obj = DatingCentury(century=century)
    session.add(obj)
    session.flush()
    cache[century] = obj
    return obj.id


def _get_or_create_vernacular_region(
    session: Session, region: Optional[str], cache: Dict[str, VernacularRegion]
) -> Optional[int]:
    raw = clean_value(region)
    if raw is None:
        return None
    if raw in cache:
        return cache[raw].id
    existing = session.exec(
        select(VernacularRegion).where(VernacularRegion.region == raw)
    ).first()
    if existing:
        cache[raw] = existing
        return existing.id
    obj = VernacularRegion(region=raw)
    session.add(obj)
    session.flush()
    cache[raw] = obj
    return obj.id


# ---------------------------------------------------------------------------
# ImportReport
# ---------------------------------------------------------------------------

class ImportReport:
    """Collects import anomalies and writes them to an Excel workbook."""

    CATEGORIES = [
        "url_skipped",
        "link_unresolved",
        "invalid_format",
        "critical_error",
        "image_warning",
    ]

    def __init__(self) -> None:
        self.categories: Dict[str, List[Dict[str, Any]]] = {
            c: [] for c in self.CATEGORIES
        }

    def add(self, category: str, data: Dict[str, Any]) -> None:
        self.categories.setdefault(category, []).append(data)

    def count(self, category: str) -> int:
        return len(self.categories.get(category, []))

    def save(self, path: str) -> None:
        from openpyxl import Workbook as WB

        wb = WB()
        if wb.active:
            wb.remove(wb.active)
        for category in self.CATEGORIES:
            entries = self.categories.get(category, [])
            ws = wb.create_sheet(title=category.upper()[:31])
            if not entries:
                ws.append(["(no entries)"])
                continue
            headers = list(entries[0].keys())
            ws.append(headers)
            for entry in sorted(entries, key=lambda x: x.get("Row", 0)):
                ws.append([entry.get(h, "") for h in headers])
        wb.save(path)


# ---------------------------------------------------------------------------
# Step 1 — import_texts
# ---------------------------------------------------------------------------

TEXTS_EXPECTED_HEADERS = [
    "BHL or NO BHL",
    "Unique identifier",
    "Title of the work",
    "Approximate token count",
    "Prose or verse",
    "Source type",
    "Subtype",
    "Réécriture?",
    "Réécriture of which text(s)?",
    "Quarter century chronology",
    "Dating range (beginning)",
    "Dating range (end)",
    "Dating notes",
    "Dating confidence rating",
    "Author of the text",
    "Is author based in destinatary institution?",
    "Institutional training ground of the author",
    "Regional or local antecedents of the author",
    "Author milieu",
    "Text creation - location by archdiocese",
    "Text creation - location by diocese",
    "Text creation - location by institution",
    "Text creation - institution - most precise possible GPS Longitude",
    "Text creation - institution - most precise possible GPS Latitude",
    "Precise institutional origin?",
    "Primary institutional destinatary",
    "Institutional destinatary - most precise possible GPS Longitude",
    "Institutional destinatary - most precise possible GPS Latitude",
    "Precise destinatary?",
    "Selected reference",
    "Notes",
]


def import_texts(session: Session, wb, report: ImportReport) -> Dict[str, "Text"]:
    """Import the 'TEXTS' sheet.

    Returns a dict mapping the text key (the 'Unique identifier' value) →
    Text instance, so manuscripts/editions can attach to it.
    """
    ws = _get_sheet(wb, SHEET_TEXTS)
    headers, rows_iter = _read_headers(ws)
    _require_headers(headers, TEXTS_EXPECTED_HEADERS, SHEET_TEXTS)

    text_cache: Dict[str, Text] = {}
    place_cache: Dict[str, Place] = {}
    church_cache: Dict[str, ChurchEntity] = {}
    author_cache: Dict[str, Author] = {}
    typology_cache: Dict[str, Typology] = {}
    anon_seq = count(1)

    inserted = skipped = 0

    logger.info("[Texts] Reading TEXTS sheet…")
    for excel_row, cells in _iter_data_rows(rows_iter, SHEET_TEXTS):
        row = Row(cells, headers)

        seen = inserted + skipped
        if seen and seen % 200 == 0:
            logger.info(f"[Texts] {seen} rows read…")

        uid = row.i("Unique identifier")
        uid_str = row.s("Unique identifier")  # used as fallback string key
        # The text key the other sheets join on is the 'Unique identifier'
        # integer (e.g. 29).
        key = _normalize_ref(uid if uid is not None else uid_str)
        if key is None:
            skipped += 1
            continue
        if key in text_cache:
            skipped += 1
            continue

        orig_arch_id = _get_or_create_church_entity(
            session, row.s("Text creation - location by archdiocese"),
            "archdiocese", church_cache,
        )
        # The source columns are mislabeled: the "GPS Longitude" column holds
        # latitude (~49 for Metz) and "GPS Latitude" holds longitude (~6).
        # Swap on read so Place.lat/lon are geographically correct.
        orig_lat = parse_gps_int(row.raw(
            "Text creation - institution - most precise possible GPS Longitude"
        ))
        orig_lon = parse_gps_int(row.raw(
            "Text creation - institution - most precise possible GPS Latitude"
        ))
        orig_place_id = _get_or_create_place(
            session, row.s("Text creation - location by institution"),
            place_cache, lat=orig_lat, lon=orig_lon,
        )
        # GIS fallback: without an origin institution, the sheet's GPS belongs
        # to the origin diocese — store it there so the map can fall back.
        orig_dio_id = _get_or_create_church_entity(
            session, row.s("Text creation - location by diocese"),
            "diocese", church_cache,
            lat=orig_lat if orig_place_id is None else None,
            lon=orig_lon if orig_place_id is None else None,
        )

        dest_lat = parse_gps_int(row.raw(
            "Institutional destinatary - most precise possible GPS Longitude"
        ))
        dest_lon = parse_gps_int(row.raw(
            "Institutional destinatary - most precise possible GPS Latitude"
        ))
        dest_place_id = _get_or_create_place(
            session, row.s("Primary institutional destinatary"),
            place_cache, lat=dest_lat, lon=dest_lon,
        )

        # '?' in the author name marks uncertain attribution: strip it from
        # the name and record it as a per-text authorship confidence rating.
        author_name, author_uncertain = _clean_uncertain(
            row.s("Author of the text")
        )
        auth_id = _get_or_create_author(
            session, author_name, author_cache, anon_seq,
            institutional_training_ground=row.s(
                "Institutional training ground of the author"
            ),
            regional_antecedents=row.s(
                "Regional or local antecedents of the author"
            ),
            milieu=row.s("Author milieu"),
        )

        src_typo_id = _get_or_create_typology(
            session, row.s("Source type"), None, typology_cache
        )
        subtype_id = _get_or_create_typology(
            session, row.s("Subtype"), src_typo_id, typology_cache
        )

        text = Text(
            bhl_or_no_bhl=row.s("BHL or NO BHL"),
            unique_identifier=uid,
            title=row.s("Title of the work"),
            approximate_token_count=row.i("Approximate token count"),
            prose_or_verse=row.s("Prose or verse"),
            quarter_century_chronology=row.s("Quarter century chronology"),
            dating_range_start=row.i("Dating range (beginning)"),
            dating_range_end=row.i("Dating range (end)"),
            dating_notes=row.s("Dating notes"),
            dating_confidence_rating=_confidence_value(
                row.raw("Dating confidence rating"),
                report, excel_row, "Dating confidence rating",
            ),
            origin_archdiocese_id=orig_arch_id,
            origin_diocese_id=orig_dio_id,
            origin_place_id=orig_place_id,
            is_origin_precise=row.yesno("Precise institutional origin?"),
            is_destinatary_precise=row.yesno("Precise destinatary?"),
            primary_destinatary_place_id=dest_place_id,
            author_id=auth_id,
            authorship_confidence_rating="uncertain" if author_uncertain else None,
            author_locally_based=row.s(
                "Is author based in destinatary institution?"
            ),
            source_type_id=src_typo_id,
            subtype_id=subtype_id,
            is_rewrite=row.yesno("Réécriture?"),
            rewrite_notes=row.s("Réécriture of which text(s)?"),
            selected_reference=row.s("Selected reference"),
            notes=row.s("Notes"),
        )
        session.add(text)
        session.flush()
        text_cache[key] = text
        inserted += 1

    session.commit()
    logger.info(f"[Texts] {inserted} inserted, {skipped} skipped.")
    return text_cache


# ---------------------------------------------------------------------------
# Step 2 — import_manuscripts
# ---------------------------------------------------------------------------

MANUSCRIPTS_EXPECTED_HEADERS = [
    "Unique text identifier",
    "Manuscript copy unique identifier per text",
    "Codex number in database",
    "Codex unique identifier",
    "Codex with multiple manuscript copies of texts from corpus",
    "Codex features n manuscript copies of texts from corpus",
    "Preservation status of manuscript copy",
    "Manuscript location",
    "Manuscript holding institution",
    "Manuscript shelfmark",
    "Folio or page range",
    "Manuscript height",
    "Manuscript width",
    "Manuscript dating by (earliest) century",
    "Manuscript dating range start",
    "Manuscript dating range end",
    "Preferred secondary reference for manuscript dating",
    "Confidence rating for manuscript dating",
    "Usable Légendiers entry for codex contents",
    "Composite?",
    "Légendiers entry link",
    "Légendiers entry code",
    "Viable alternative for Légendiers entry on codex contents",
    "Notes on codex contents",
    "Online catalogue link",
    "Bollandist catalogue link",
    "Other relevant catalogue link",
    "Type of online images",
    "Online manuscript images",
    "Vernacular region (Romance/Germanic)",
    "Manuscript origin by archdiocese",
    "Manuscript origin by diocese",
    "Manuscript origin by diocese confidence rating",
    "Manuscript origin by diocese GPS Longitude",
    "Manuscript origin by diocese GPS latitude",
    "Manuscript origin by institution",
    "Manuscript origin confidence rating",
    "Manuscript provenance by early/earliest institutional owner",
    "Manuscript provenance by early/earliest institutional owner confidence rating",
    # The duplicate-headered columns holding the actual provenance owner
    # name/confidence (the first pair holds GPS — workbook header defect).
    "Manuscript provenance by early/earliest institutional owner_2",
    "Manuscript provenance by early/earliest institutional owner confidence rating_2",
    "Manuscript provenance by early/earliest institutional owner GPS Longitude",
    "Manuscript provenance by early/earliest institutional owner GPS latitude",
    "Manuscript provenance by undetermined or later institutional owner",
    "Manuscript provenance by undetermined or later institutional owner GPS Longitude",
    "Manuscript provenance by undetermined or later institutional owner GPS latitude",
    "Manuscript origin and provenance preferred secondary reference",
    "Based on exemplar",
    "Exemplar of which manuscript(s)",
    "Manuscript type",
    "Notes",
]


def import_manuscripts(
    session: Session,
    wb,
    text_cache: Dict[str, "Text"],
    report: ImportReport,
) -> Tuple[Dict[str, Manuscript], Dict[str, Manuscript]]:
    """Import the 'MANUSCRIPTS' sheet: one Codex per codex identifier,
    one Manuscript copy per sheet row.

    Returns:
        ms_by_number — normalised copy identifier ('29-1') → Manuscript copy
        codex_copies — normalised codex identifier → list of its copies
                       (used by editions to resolve codex-style references)
    """
    ws = _get_sheet(wb, SHEET_MANUSCRIPTS)
    headers, rows_iter = _read_headers(ws)
    _require_headers(headers, MANUSCRIPTS_EXPECTED_HEADERS, SHEET_MANUSCRIPTS)

    place_cache: Dict[str, Place] = {}
    inst_cache: Dict[str, Institution] = {}
    church_cache: Dict[str, ChurchEntity] = {}
    mt_cache: Dict[str, ManuscriptType] = {}
    century_cache: Dict[int, DatingCentury] = {}
    vernacular_cache: Dict[str, VernacularRegion] = {}
    image_type_cache: Dict[str, ImageType] = {}

    stats: Dict[str, int] = {
        "codices": 0, "copies": 0, "skipped": 0, "images": 0,
    }

    codex_by_norm: Dict[str, Codex] = {}
    codex_copies: Dict[str, List[Manuscript]] = {}
    ms_by_number: Dict[str, Manuscript] = {}

    # Deferred relation tasks:
    # (src_copy_norm, tgt_copy_norm, relation_type, certainty, notes, col, row)
    relation_tasks: List[Tuple[str, str, RelationType, Certainty,
                               Optional[str], str, int]] = []

    logger.info("[Manuscripts] Reading MANUSCRIPTS sheet…")
    batch_no = 0
    for batch in _chunked(_iter_data_rows(rows_iter, SHEET_MANUSCRIPTS), 500):
        batch_no += 1
        try:
            for excel_row, cells in batch:
                row = Row(cells, headers)

                coll_id = row.s("Codex unique identifier")
                ms_number = row.s("Manuscript copy unique identifier per text")
                text_key = _normalize_ref(row.raw("Unique text identifier"))

                if coll_id is None and ms_number is None:
                    continue

                coll_norm = _normalize_ref(coll_id)
                num_norm = _normalize_ref(ms_number)

                # --- Codex (created on first sight of its identifier) ---
                codex = codex_by_norm.get(coll_norm) if coll_norm else None
                if codex is None and coll_norm:
                    coll_place_id = _get_or_create_place(
                        session, row.s("Manuscript location"), place_cache
                    )
                    heri_inst_id = _get_or_create_institution(
                        session, row.s("Manuscript holding institution"),
                        coll_place_id, inst_cache,
                    )
                    century_id = _get_or_create_dating_century(
                        session, row.raw("Manuscript dating by (earliest) century"),
                        century_cache, report, excel_row,
                    )
                    vernacular_id = _get_or_create_vernacular_region(
                        session, row.s("Vernacular region (Romance/Germanic)"),
                        vernacular_cache,
                    )

                    origin_arch_id = _get_or_create_church_entity(
                        session, row.s("Manuscript origin by archdiocese"),
                        "archdiocese", church_cache,
                    )
                    # Diocese GPS (mislabeled lon/lat — see import_texts: swap).
                    # These are the DIOCESE coordinates, stored on the
                    # ChurchEntity itself.
                    dio_lat = parse_gps_int(row.raw("Manuscript origin by diocese GPS Longitude"))
                    dio_lon = parse_gps_int(row.raw("Manuscript origin by diocese GPS latitude"))
                    origin_dio_id = _get_or_create_church_entity(
                        session, row.s("Manuscript origin by diocese"),
                        "diocese", church_cache, lat=dio_lat, lon=dio_lon,
                    )
                    # WORKBOOK HEADER DEFECT: the two columns after
                    # 'Manuscript origin confidence rating' are labeled
                    # 'Manuscript provenance by early/earliest institutional
                    # owner' (+ '…confidence rating') but actually contain the
                    # ORIGIN INSTITUTION's GPS latitude / longitude as decimal
                    # degrees.  The real provenance owner name/confidence live
                    # in the following duplicate-headered columns, addressable
                    # via the '_2' suffix that _read_headers assigns.
                    inst_lat = parse_gps_int(row.raw(
                        "Manuscript provenance by early/earliest institutional owner"
                    ))
                    inst_lon = parse_gps_int(row.raw(
                        "Manuscript provenance by early/earliest institutional owner confidence rating"
                    ))
                    origin_place_id = _get_or_create_place(
                        session, row.s("Manuscript origin by institution"),
                        place_cache, lat=inst_lat, lon=inst_lon,
                    )

                    # Provenance owner: name/confidence from the '_2' columns;
                    # its GPS columns hold latitude in '…GPS Longitude' and
                    # longitude in '…GPS latitude' (same swap as elsewhere).
                    prov_lat = parse_gps_int(row.raw(
                        "Manuscript provenance by early/earliest institutional owner GPS Longitude"
                    ))
                    prov_lon = parse_gps_int(row.raw(
                        "Manuscript provenance by early/earliest institutional owner GPS latitude"
                    ))
                    prov_inst_id = _get_or_create_institution(
                        session,
                        row.s("Manuscript provenance by early/earliest institutional owner_2"),
                        None, inst_cache, lat=prov_lat, lon=prov_lon,
                    )
                    later_lat = parse_gps_int(row.raw(
                        "Manuscript provenance by undetermined or later institutional owner GPS Longitude"
                    ))
                    later_lon = parse_gps_int(row.raw(
                        "Manuscript provenance by undetermined or later institutional owner GPS latitude"
                    ))
                    prov_later_id = _get_or_create_institution(
                        session,
                        row.s("Manuscript provenance by undetermined or later institutional owner"),
                        None, inst_cache, lat=later_lat, lon=later_lon,
                    )
                    ms_typo_id = _get_or_create_manuscript_type(
                        session, row.s("Manuscript type"), mt_cache
                    )

                    codex = Codex(
                        codex_unique_identifier=coll_id,
                        codex_number_in_database=row.i("Codex number in database"),
                        codex_with_multiple_copies=row.yesno(
                            "Codex with multiple manuscript copies of texts from corpus"
                        ),
                        codex_copies_count=row.i(
                            "Codex features n manuscript copies of texts from corpus"
                        ),
                        is_composite_codex=row.yesno("Composite?"),
                        location_place_id=coll_place_id,
                        holding_institution_id=heri_inst_id,
                        shelfmark=row.s("Manuscript shelfmark"),
                        dating_century_id=century_id,
                        dating_range_start=row.i("Manuscript dating range start"),
                        dating_range_end=row.i("Manuscript dating range end"),
                        dating_reference=row.s("Preferred secondary reference for manuscript dating"),
                        dating_confidence=_confidence_value(
                            row.raw("Confidence rating for manuscript dating"),
                            report, excel_row, "Confidence rating for manuscript dating",
                        ),
                        legendiers_usable=row.yesno("Usable Légendiers entry for codex contents"),
                        legendiers_code=row.s("Légendiers entry code"),
                        legendiers_alternative=row.s(
                            "Viable alternative for Légendiers entry on codex contents"
                        ),
                        legendiers_notes=row.s("Notes on codex contents"),
                        origin_archdiocese_id=origin_arch_id,
                        origin_diocese_id=origin_dio_id,
                        origin_diocese_confidence=_confidence_value(
                            row.raw("Manuscript origin by diocese confidence rating"),
                            report, excel_row, "Manuscript origin by diocese confidence rating",
                        ),
                        origin_place_id=origin_place_id,
                        origin_confidence=_confidence_value(
                            row.raw("Manuscript origin confidence rating"),
                            report, excel_row, "Manuscript origin confidence rating",
                        ),
                        provenance_institution_id=prov_inst_id,
                        # '_2': the first-occurrence column of this header
                        # actually holds GPS (see comment above).
                        provenance_institution_confidence=_confidence_value(
                            row.raw("Manuscript provenance by early/earliest institutional owner confidence rating_2"),
                            report, excel_row,
                            "Manuscript provenance by early/earliest institutional owner confidence rating_2",
                        ),
                        provenance_later_institution_id=prov_later_id,
                        provenance_reference=row.s("Manuscript origin and provenance preferred secondary reference"),
                        vernacular_region_id=vernacular_id,
                        manuscript_type_id=ms_typo_id,
                        dimension_width_cm=row.f("Manuscript width"),
                        dimension_height_cm=row.f("Manuscript height"),
                    )
                    session.add(codex)
                    session.flush()
                    codex_by_norm[coll_norm] = codex
                    codex_copies[coll_norm] = []
                    stats["codices"] += 1

                    # Légendiers entry link (hyperlink target)
                    leg_url = _extract_hyperlink_url(row.cell("Légendiers entry link"))
                    if leg_url:
                        codex.legendiers_link = leg_url

                    # Catalogue links → ExternalResource.  The three columns
                    # may repeat the same URL; dedupe to respect the
                    # (codex_id, url) unique constraint.
                    seen_urls: set = set()
                    for col, rtype in (
                        ("Online catalogue link", "catalog_link"),
                        ("Bollandist catalogue link", "bollandist_catalog"),
                        ("Other relevant catalogue link", "catalog_link"),
                    ):
                        url = _extract_hyperlink_url(row.cell(col))
                        if url and url not in seen_urls:
                            seen_urls.add(url)
                            session.add(ExternalResource(
                                codex_id=codex.id, url=url,
                                resource_type=_resource_type(rtype),
                            ))

                    # Images
                    img_avail = row.s("Type of online images")
                    img_url = _extract_hyperlink_url(row.cell("Online manuscript images"))
                    if img_url:
                        itype_id = _get_or_create_image_type(
                            session, _infer_image_type(img_avail), image_type_cache
                        )
                        session.add(Image(
                            url=img_url, image_type_id=itype_id, codex_id=codex.id
                        ))
                        stats["images"] += 1
                    elif img_avail:
                        report.add("image_warning", {
                            "Row": excel_row,
                            "ImageAvailability": img_avail,
                            "Reason": "Type-of-images set but no image URL found",
                        })

                # --- Manuscript copy (one per sheet row) ---
                if ms_number is None:
                    report.add("invalid_format", {
                        "Row": excel_row,
                        "Column": "Manuscript copy unique identifier per text",
                        "Value": "",
                        "Reason": "Row has no copy identifier — no copy created",
                    })
                    stats["skipped"] += 1
                    continue
                if num_norm in ms_by_number:
                    report.add("invalid_format", {
                        "Row": excel_row,
                        "Column": "Manuscript copy unique identifier per text",
                        "Value": ms_number,
                        "Reason": "Duplicate copy identifier — row skipped",
                    })
                    stats["skipped"] += 1
                    continue

                text_obj = text_cache.get(text_key) if text_key else None
                if text_key and text_obj is None:
                    report.add("link_unresolved", {
                        "Row": excel_row,
                        "Column": "Unique text identifier",
                        "Source": ms_number,
                        "Target": str(text_key),
                        "Reason": "Copy references a text not in TEXTS sheet",
                    })

                copy = Manuscript(
                    manuscript_copy_identifier_per_text=ms_number,
                    codex_id=codex.id if codex else None,
                    text_id=text_obj.id if text_obj else None,
                    preservation_status=row.s("Preservation status of manuscript copy"),
                    folio_or_page_range=row.s("Folio or page range"),
                    notes=row.s("Notes"),
                )
                session.add(copy)
                session.flush()
                ms_by_number[num_norm] = copy
                if coll_norm:
                    codex_copies[coll_norm].append(copy)
                stats["copies"] += 1

                # --- Exemplar relations (resolved in post-pass) ---
                # 'Based on exemplar': this copy is a copy_of the target.
                based_on = row.s("Based on exemplar")
                if based_on:
                    for tgt in re.split(r"[;,]", based_on):
                        tgt_norm = _normalize_ref(tgt)
                        if tgt_norm:
                            relation_tasks.append((
                                num_norm, tgt_norm,
                                RelationType.copy_of,
                                Certainty.uncertain if "?" in tgt else Certainty.certain,
                                row.s("Notes"), "Based on exemplar", excel_row,
                            ))
                # 'Exemplar of which manuscript(s)': this copy is the
                # exemplar_of the listed copies.
                exemplar_of = row.s("Exemplar of which manuscript(s)")
                if exemplar_of:
                    for child in re.split(r"[;,]", exemplar_of):
                        child_norm = _normalize_ref(child)
                        if child_norm:
                            relation_tasks.append((
                                num_norm, child_norm,
                                RelationType.exemplar_of,
                                Certainty.uncertain if "?" in child else Certainty.certain,
                                row.s("Notes"), "Exemplar of which manuscript(s)",
                                excel_row,
                            ))

            session.commit()
            logger.info(
                f"[Manuscripts] batch {batch_no} committed "
                f"({stats['copies']} copies so far)…"
            )
        except Exception as e:
            logger.error(f"[Manuscripts] Batch error: {e}")
            report.add("critical_error", {"Error": str(e)})
            session.rollback()
            continue

    _resolve_relations(session, relation_tasks, ms_by_number, report)

    logger.info(
        f"[Manuscripts] {stats['codices']} codices, {stats['copies']} copies, "
        f"{stats['skipped']} rows skipped, {stats['images']} images."
    )
    return ms_by_number, codex_copies


def _resolve_relations(
    session: Session,
    tasks: List[Tuple[str, str, RelationType, Certainty, Optional[str], str, int]],
    ms_by_number: Dict[str, Manuscript],
    report: ImportReport,
) -> None:
    """Resolve deferred exemplar/copy relations, verifying both endpoints."""
    logger.info(f"[Relations] Resolving {len(tasks)} relation tasks...")
    created = 0
    for src_num, tgt_num, rel_type, certainty, notes, col, excel_row in tasks:
        src = ms_by_number.get(src_num)
        tgt = ms_by_number.get(tgt_num)
        if src is None or tgt is None:
            report.add("link_unresolved", {
                "Row": excel_row,
                "Column": col,
                "Source": src_num,
                "Target": tgt_num,
                "Reason": (
                    f"Unresolved copy identifier "
                    f"({'source' if src is None else 'target'} not found)"
                ),
            })
            continue
        exists = session.exec(
            select(ManuscriptRelation).where(
                ManuscriptRelation.source_manuscript_id == src.id,
                ManuscriptRelation.target_manuscript_id == tgt.id,
                ManuscriptRelation.relation_type == rel_type,
            )
        ).first()
        if exists:
            continue
        session.add(ManuscriptRelation(
            source_manuscript_id=src.id,
            target_manuscript_id=tgt.id,
            relation_type=rel_type,
            certainty=certainty,
            notes=notes,
            source_reference=f"{SHEET_MANUSCRIPTS} (row {excel_row}, '{col}')",
        ))
        created += 1
    session.commit()
    logger.info(f"[Relations] {created} relations created.")


def _resource_type(name: str) -> ExternalResourceType:
    try:
        return ExternalResourceType(name)
    except ValueError:
        return ExternalResourceType.other


# ---------------------------------------------------------------------------
# Step 3 — import_editions
# ---------------------------------------------------------------------------

EDITIONS_EXPECTED_HEADERS = [
    "Unique identifier",
    "Title",
    "Edition unique identifier per individual text",
    "Edition unique identifier (inc. volume)",
    "Publication year",
    "Edition reference",
    "Page numbers",
    "Reprint ?",
    "If reprint, identically typeset?",
    "If reprint, newly typeset?",
    "If reprint, of what?",
    "Images of edition?",
    "Edition images link",
    "Transcription available?",
    "Collation done?",
    "Notes",
] + [f"Manuscript used {n}" for n in range(1, 17)] \
  + [f"Edition used or consulted {n}" for n in range(1, 6)]


def import_editions(
    session: Session,
    wb,
    text_cache: Dict[str, "Text"],
    ms_by_number: Dict[str, Manuscript],
    codex_copies: Dict[str, List[Manuscript]],
    report: ImportReport,
) -> None:
    """Import the 'EDITIONS' sheet.

    Each edition links to its text via 'Unique identifier' and to the
    manuscript copies it used via the 'Manuscript used 1..16' columns.  Those
    cells hold either copy identifiers ('29-1', resolved via ms_by_number) or
    codex identifiers ('Cologne HA 6' — resolved to that codex's copies of
    the edition's text); genuine placeholders ('to be verified',
    'Unpublished') are reported as unresolved.  The paired 'Likely use of a
    copy of Manuscript N?' cells become EditionManuscript.likely_copy.

    'Edition unique identifier (inc. volume)' becomes an EditionVolume row;
    'Edition used or consulted 1..5' values are matched against the volumes
    in a deferred pass → EditionConsultedVolume.

    Rows are keyed on 'Edition unique identifier per individual text'
    (e.g. "693-B") — one edition row per text-edition.
    """
    ws = _get_sheet(wb, SHEET_EDITIONS)
    headers, rows_iter = _read_headers(ws)
    _require_headers(headers, EDITIONS_EXPECTED_HEADERS, SHEET_EDITIONS)

    resource_cache: Dict[str, ExternalResource] = {}
    volume_cache: Dict[str, EditionVolume] = {}
    stats = {"inserted": 0, "ms_links": 0, "scan_links": 0, "consulted_links": 0}

    ms_used_pairs = [
        (f"Manuscript used {n}", f"Likely use of a copy of Manuscript {n}?")
        for n in range(1, 17)
    ]
    consulted_cols = [f"Edition used or consulted {n}" for n in range(1, 6)]

    # Deferred: consulted refs may name volumes that appear later in the sheet.
    consulted_tasks: List[Tuple[Edition, List[str], int]] = []

    logger.info("[Editions] Reading EDITIONS sheet…")
    batch_no = 0
    for batch in _chunked(_iter_data_rows(rows_iter, SHEET_EDITIONS), 500):
        batch_no += 1
        try:
            for excel_row, cells in batch:
                row = Row(cells, headers)

                text_key = _normalize_ref(row.raw("Unique identifier"))
                uid_desc = row.s("Edition unique identifier (inc. volume)")
                ed_ref = row.s("Edition unique identifier per individual text")
                if uid_desc is None and ed_ref is None:
                    continue

                text_obj = text_cache.get(text_key) if text_key else None
                if text_key and text_obj is None:
                    report.add("link_unresolved", {
                        "Row": excel_row,
                        "Column": "Unique identifier",
                        "Source": text_key,
                        "Target": "",
                        "Reason": "Edition references a text not in TEXTS sheet",
                    })

                consulted = [row.s(c) for c in consulted_cols]
                consulted = [c for c in consulted if c]

                volume = _get_or_create_edition_volume(
                    session, uid_desc, volume_cache
                )

                # One edition row per text-edition: key on the per-text
                # edition identifier.
                existing = None
                if ed_ref:
                    existing = session.exec(
                        select(Edition).where(
                            Edition.edition_unique_identifier_per_text == ed_ref
                        )
                    ).first()
                if existing:
                    edition = existing
                else:
                    edition = Edition(
                        title=row.s("Title"),
                        edition_unique_identifier_per_text=ed_ref,
                        text_unique_identifier=parse_int(row.raw("Unique identifier")),
                        volume_id=volume.id if volume else None,
                        publication_year=row.i("Publication year"),
                        edition_reference=row.s("Edition reference"),
                        page_numbers=row.s("Page numbers"),
                        is_reprint=row.yesno("Reprint ?"),
                        reprint_identically_typeset=row.yesno(
                            "If reprint, identically typeset?"
                        ),
                        reprint_newly_typeset=row.yesno("If reprint, newly typeset?"),
                        reprint_of=row.s("If reprint, of what?"),
                        images_of_edition=row.s("Images of edition?"),
                        edition_images_link=_extract_hyperlink_url(
                            row.cell("Edition images link")
                        ),
                        transcription_available=row.yesno("Transcription available?"),
                        collation_done=row.yesno("Collation done?"),
                        editions_consulted=", ".join(consulted) or None,
                        notes=row.s("Notes"),
                        text_id=text_obj.id if text_obj else None,
                    )
                    session.add(edition)
                    session.flush()
                    stats["inserted"] += 1
                    if consulted:
                        consulted_tasks.append((edition, consulted, excel_row))

                # Scan / edition image link
                scan_url = _extract_hyperlink_url(row.cell("Edition images link"))
                if scan_url:
                    resource = resource_cache.get(scan_url)
                    if resource is None:
                        resource = session.exec(
                            select(ExternalResource).where(
                                ExternalResource.url == scan_url,
                                ExternalResource.codex_id == None,  # noqa: E711
                            )
                        ).first()
                    if resource is None:
                        resource = ExternalResource(
                            url=scan_url, resource_type=ExternalResourceType.scan
                        )
                        session.add(resource)
                        session.flush()
                    resource_cache[scan_url] = resource
                    link_exists = session.exec(
                        select(EditionExternalResource).where(
                            EditionExternalResource.edition_id == edition.id,
                            EditionExternalResource.resource_id == resource.id,
                        )
                    ).first()
                    if not link_exists:
                        session.add(EditionExternalResource(
                            edition_id=edition.id, resource_id=resource.id
                        ))
                        stats["scan_links"] += 1

                # Manuscripts used → EditionManuscript, with FK verification.
                for col, likely_col in ms_used_pairs:
                    ms_ref = row.s(col)
                    if ms_ref is None:
                        continue
                    likely = row.yesno(likely_col)
                    ms_norm = _normalize_ref(ms_ref)

                    # Copy identifier first; codex identifier resolves to the
                    # codex's copies of this edition's text.
                    targets: List[Manuscript] = []
                    copy = ms_by_number.get(ms_norm)
                    if copy is not None:
                        targets = [copy]
                    elif ms_norm in codex_copies:
                        candidates = codex_copies[ms_norm]
                        if text_obj is not None:
                            targets = [c for c in candidates
                                       if c.text_id == text_obj.id]
                        if not targets:
                            report.add("link_unresolved", {
                                "Row": excel_row,
                                "Column": col,
                                "Source": uid_desc or ed_ref,
                                "Target": ms_ref,
                                "Reason": "Codex found but has no copy of this "
                                          "edition's text",
                            })
                            continue
                    else:
                        report.add("link_unresolved", {
                            "Row": excel_row,
                            "Column": col,
                            "Source": uid_desc or ed_ref,
                            "Target": ms_ref,
                            "Reason": "No copy or codex with this identifier "
                                      "in MANUSCRIPTS sheet",
                        })
                        continue

                    for target in targets:
                        exists = session.exec(
                            select(EditionManuscript).where(
                                EditionManuscript.edition_id == edition.id,
                                EditionManuscript.ms_id == target.id,
                            )
                        ).first()
                        if not exists:
                            session.add(EditionManuscript(
                                edition_id=edition.id, ms_id=target.id,
                                inspection_status="unknown",
                                likely_copy=likely,
                            ))
                            stats["ms_links"] += 1

            session.commit()
            logger.info(
                f"[Editions] batch {batch_no} committed "
                f"({stats['inserted']} editions so far)…"
            )
        except Exception as e:
            logger.error(f"[Editions] Batch error: {e}")
            report.add("critical_error", {"Error": str(e)})
            session.rollback()
            continue

    # --- Deferred pass: consulted volumes ---------------------------------
    seen_pairs: set = set()
    for edition, refs, excel_row in consulted_tasks:
        for ref in refs:
            vol = volume_cache.get(_normalize_vol(ref))
            if vol is None:
                report.add("link_unresolved", {
                    "Row": excel_row,
                    "Column": "Edition used or consulted",
                    "Source": edition.edition_unique_identifier_per_text or "",
                    "Target": ref,
                    "Reason": "Consulted reference matches no edition volume",
                })
                continue
            pair = (edition.id, vol.id)
            if pair in seen_pairs:
                continue
            seen_pairs.add(pair)
            session.add(EditionConsultedVolume(
                edition_id=edition.id, volume_id=vol.id
            ))
            stats["consulted_links"] += 1
    session.commit()

    logger.info(
        f"[Editions] {stats['inserted']} editions, "
        f"{stats['ms_links']} manuscript links, {stats['scan_links']} scan links, "
        f"{stats['consulted_links']} consulted-volume links."
    )


# ---------------------------------------------------------------------------
# Post-import validation
# ---------------------------------------------------------------------------

_FLOAT_RE = re.compile(r"^-?\d+(?:[.,]\d+)?$")

# Name columns where a bare number means a column was mis-mapped
# (e.g. GPS coordinates imported as institution names).
_NAME_SANITY_COLUMNS = [
    ("place", "name"),
    ("institution", "name"),
    ("churchentity", "name"),
    ("author", "name"),
]


def validate_import(session: Session, report: ImportReport) -> int:
    """Belt-and-braces integrity check after the import.

    1. Referential integrity: every declared FK is scanned for orphans.
       PostgreSQL already enforces these constraints; a non-zero count here
       means the schema lost a FK declaration.
    2. Data sanity: name and confidence-rating columns must not contain bare
       numbers — that pattern indicates a mis-mapped source column (this is
       exactly how GPS-in-provenance was discovered).

    Returns the number of problems found (also logged + reported).
    """
    from sqlalchemy import text as sa_text

    conn = session.connection()
    problems = 0

    logger.info("[Validate] Checking foreign-key integrity…")
    fk_pairs = 0
    for table in SQLModel.metadata.tables.values():
        for fk in sorted(table.foreign_keys, key=lambda f: f.parent.name):
            col, ref = fk.parent, fk.column
            fk_pairs += 1
            orphans = conn.execute(sa_text(
                f'SELECT count(*) FROM "{table.name}" t '
                f'LEFT JOIN "{ref.table.name}" r ON t."{col.name}" = r."{ref.name}" '
                f'WHERE t."{col.name}" IS NOT NULL AND r."{ref.name}" IS NULL'
            )).scalar_one()
            if orphans:
                problems += orphans
                logger.error(
                    f"[Validate] {orphans} orphaned {table.name}.{col.name} "
                    f"→ {ref.table.name}.{ref.name}"
                )
                report.add("critical_error", {
                    "Error": f"{orphans} orphaned FK values in "
                             f"{table.name}.{col.name}",
                })

    logger.info(f"[Validate] {fk_pairs} FK relationships checked.")

    confidence_cols = [
        (t.name, c.name)
        for t in SQLModel.metadata.tables.values()
        for c in t.columns if "confidence" in c.name
    ]
    for tname, cname in _NAME_SANITY_COLUMNS + confidence_cols:
        rows = conn.execute(sa_text(
            f'SELECT DISTINCT "{cname}" FROM "{tname}" WHERE "{cname}" IS NOT NULL'
        )).all()
        bad = [r[0] for r in rows
               if isinstance(r[0], str) and _FLOAT_RE.match(r[0].strip())]
        if bad:
            problems += len(bad)
            logger.error(
                f"[Validate] {tname}.{cname} contains numeric-looking values "
                f"(mis-mapped source column?): {bad[:5]}"
            )
            for v in bad:
                report.add("invalid_format", {
                    "Row": 0,
                    "Column": f"{tname}.{cname}",
                    "Value": v,
                    "Reason": "Numeric value in a name/confidence column — "
                              "suspected mis-mapped source column",
                })

    if problems:
        logger.error(f"[Validate] FAILED: {problems} problem(s) found "
                     f"({fk_pairs} FK pairs scanned).")
    else:
        logger.info(f"[Validate] OK: {fk_pairs} FK pairs scanned, 0 orphans; "
                    "name/confidence columns clean.")
    return problems


# ---------------------------------------------------------------------------
# Entrypoint
# ---------------------------------------------------------------------------

def main() -> None:
    """Run the full Excel → PostgreSQL import pipeline."""
    logger.info("Importing into PostgreSQL")
    if not EXCEL.exists():
        logger.error(f"Excel file not found at {EXCEL}, cannot proceed.")
        return

    SQLModel.metadata.create_all(engine)

    wb = load_workbook(EXCEL, data_only=True)
    report = ImportReport()

    with Session(engine) as session:
        logger.info("=== Step 1: Importing Texts ===")
        text_cache = import_texts(session, wb, report)

        logger.info("=== Step 2: Importing Manuscripts ===")
        ms_by_number, codex_copies = import_manuscripts(
            session, wb, text_cache, report
        )

        logger.info("=== Step 3: Importing Editions ===")
        import_editions(session, wb, text_cache, ms_by_number, codex_copies, report)

        logger.info("=== Step 4: Validating referential integrity ===")
        validate_import(session, report)

    report.save(str(DATA_ROOT / "import_report.xlsx"))
    logger.info(
        f"[Report] link_unresolved={report.count('link_unresolved')}, "
        f"image_warning={report.count('image_warning')}, "
        f"critical_error={report.count('critical_error')}"
    )


if __name__ == "__main__":
    main()
