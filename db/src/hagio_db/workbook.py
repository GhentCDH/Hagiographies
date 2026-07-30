"""Read the corpus workbook.

Only what the backfill needs, following the conventions the importer
established in utils/importer/src/importer/excel.py: read-only + data_only
loading, sheet lookup by *stripped* name (the sheet is literally
' MANUSCRIPTS'), header row 1, whitespace-collapsing header normalisation
(several headers contain a non-breaking space).
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Iterator, NamedTuple

import openpyxl

NA_VALUES = {"", "n/a", "na", "-"}


class Row(NamedTuple):
    number: int  # 1-based Excel row number, so the report can point at it
    values: tuple


def load(path: Path):
    return openpyxl.load_workbook(path, read_only=True, data_only=True)


def sheet(workbook, name: str):
    for candidate in workbook.sheetnames:
        if candidate.strip() == name:
            return workbook[candidate]
    raise KeyError(f"sheet {name!r} not found in {workbook.sheetnames}")


def _normalise_header(header: str) -> str:
    return re.sub(r"\s+", " ", header).strip()


def headers(ws) -> dict[str, int]:
    """Normalised header -> 0-based column index (first occurrence wins)."""
    first = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
    positions: dict[str, int] = {}
    for index, header in enumerate(first):
        if header is None:
            continue
        positions.setdefault(_normalise_header(str(header)), index)
    return positions


def require(positions: dict[str, int], *wanted: str) -> None:
    missing = [w for w in wanted if _normalise_header(w) not in positions]
    if missing:
        raise KeyError(f"workbook is missing headers: {missing}")


def rows(ws) -> Iterator[Row]:
    for number, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(value is None for value in values):
            continue
        yield Row(number, values)


def norm(value) -> str | None:
    """Cell -> trimmed string, or None for blank and 'N/A'.

    Integral floats become plain integers ('29.0' -> '29'), matching how the
    importer's strict_str reads numeric cells.
    """
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    text = re.sub(r"\s+", " ", str(value)).strip()
    return None if text.lower() in NA_VALUES else text


def norm_exact(value) -> str | None:
    """Like norm(), but strip-only and without the N/A rule.

    Used for the parts identifiers are built from, so we reproduce the
    importer's strict_str byte for byte and match what it wrote.
    """
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return str(value).strip() or None


def cell(row: Row, positions: dict[str, int], header: str) -> str | None:
    return at(row, positions[_normalise_header(header)])


def cell_exact(row: Row, positions: dict[str, int], header: str) -> str | None:
    index = positions[_normalise_header(header)]
    if index >= len(row.values):
        return None
    return norm_exact(row.values[index])


def at(row: Row, index: int) -> str | None:
    if index >= len(row.values):
        return None
    return norm(row.values[index])
