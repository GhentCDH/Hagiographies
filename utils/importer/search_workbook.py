
from openpyxl import load_workbook
import os

f = '/data/hagiographies.xlsx'
if not os.path.exists(f):
    print(f"File not found: {f}")
    exit(1)

wb = load_workbook(f, read_only=True, data_only=True)
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    # Check first 50 rows for any cell containing keywords
    for i, row in enumerate(ws.iter_rows(max_row=50)):
        for j, cell in enumerate(row):
            if cell.value:
                val = str(cell.value).lower()
                if 'exemplar' in val or 'copy' in val or 'witness' in val:
                    print(f"[{sheet_name}] Row {i+1} Col {j+1}: '{cell.value}'")
