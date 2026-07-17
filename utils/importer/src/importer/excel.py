"""Workbook access: sheet lookup, header verification, row iteration.

Sheet and header names are matched on their stripped form because the
workbook contains e.g. a ' MANUSCRIPTS' sheet (leading space) and headers
with trailing spaces. A missing sheet or header is a fatal error — the
workbook layout is a contract, not something to guess around.
"""

import re
from collections.abc import Iterator
from pathlib import Path
from typing import Any
from xml.etree import ElementTree

import openpyxl
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


class WorkbookError(RuntimeError):
    """The workbook does not match the expected layout (fatal, exit 2)."""


def load(path: Path) -> Workbook:
    """Read-only load, deliberately: besides being faster, openpyxl's full
    loader materializes phantom cell values out of orphaned hyperlink records
    (the workbook has hyperlinks on rows whose contents were deleted), which
    would turn empty rows into bogus data rows. Read-only mode does not parse
    hyperlinks at all — cell_hyperlinks() reads them straight from the XML.
    """
    if not path.exists():
        raise WorkbookError(f"workbook not found: {path}")
    return openpyxl.load_workbook(path, read_only=True, data_only=True)


def sheet(workbook: Workbook, name: str) -> Worksheet:
    """Find a sheet by stripped name."""
    for candidate in workbook.sheetnames:
        if candidate.strip() == name:
            return workbook[candidate]
    raise WorkbookError(
        f"sheet {name!r} not found; workbook has {workbook.sheetnames}"
    )


def _normalize_header(header: str) -> str:
    """Collapse all whitespace runs (incl. non-breaking spaces) to one space.

    The workbook has headers like 'Likely\\xa0use of a copy of Manuscript 3?'
    where a non-breaking space stands in for a regular one.
    """
    return re.sub(r"\s+", " ", header).strip()


def header_map(ws: Worksheet, expected: list[str]) -> dict[str, int]:
    """Map each expected header to its 0-based column index.

    Headers are matched on their whitespace-normalized form (row 1). Every
    expected header must be present.
    """
    first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
    positions: dict[str, int] = {}
    for index, header in enumerate(first_row):
        if header is None:
            continue
        positions.setdefault(_normalize_header(str(header)), index)
    missing = [h for h in expected if _normalize_header(h) not in positions]
    if missing:
        raise WorkbookError(f"sheet {ws.title.strip()!r} is missing headers: {missing}")
    return {h: positions[_normalize_header(h)] for h in expected}


def header_positions(ws: Worksheet, header: str) -> list[int]:
    """All 0-based column indices whose normalized header matches.

    For genuinely duplicated headers (the MANUSCRIPTS provenance owner and
    confidence headers each appear twice); header_map always resolves to the
    first occurrence.
    """
    first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
    wanted = _normalize_header(header)
    return [
        index
        for index, cell in enumerate(first_row)
        if cell is not None and _normalize_header(str(cell)) == wanted
    ]


def data_rows(ws: Worksheet) -> Iterator[tuple[int, tuple[Any, ...]]]:
    """Yield (excel_row_number, values) for every row below the header."""
    for number, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        yield number, values


_MAIN_NS = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
_REL_ID = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"


def cell_hyperlinks(ws: Worksheet) -> dict[str, str]:
    """Map cell references (e.g. 'N12') to their hyperlink target URL.

    Read-only worksheets (see load) never parse hyperlinks, so they are read
    straight from the sheet XML and its relationships in the workbook archive
    (still open on a read-only workbook; _archive/_worksheet_path are private
    openpyxl attributes, stable across the 3.1 line we depend on).
    """
    archive = ws.parent._archive
    sheet_path = ws._worksheet_path.lstrip("/")
    directory, _, filename = sheet_path.rpartition("/")
    rels_path = f"{directory}/_rels/{filename}.rels"
    targets: dict[str, str] = {}
    if rels_path in archive.namelist():
        targets = {
            rel.get("Id"): rel.get("Target")
            for rel in ElementTree.fromstring(archive.read(rels_path))
        }
    hyperlinks = ElementTree.fromstring(archive.read(sheet_path)).find(
        f"{_MAIN_NS}hyperlinks"
    )
    links: dict[str, str] = {}
    for link in hyperlinks if hyperlinks is not None else ():
        # Links without a relationship id are internal anchors, not URLs.
        target = targets.get(link.get(_REL_ID))
        if not target:
            continue
        ref = link.get("ref")
        if ":" in ref:  # a hyperlink can span a cell range
            min_col, min_row, max_col, max_row = range_boundaries(ref)
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    links[f"{get_column_letter(col)}{row}"] = target
        else:
            links[ref] = target
    return links


def is_empty(values: tuple[Any, ...]) -> bool:
    """A row with no meaningful content in any cell."""
    return all(v is None or (isinstance(v, str) and not v.strip()) for v in values)
